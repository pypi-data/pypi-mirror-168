import time
import re
import openpyxl
import clearing
from rich.console import Console
from rich.table import Table


def search_drugs():
    clearing.clear()
    console = Console()
    table = Table(show_header=False, header_style="bold blue",
                  title="SEARCH FOR MEDICINE", title_justify="center")
    table.add_row("Search your inventory for medicines ..")
    console.print(table)
    while (True):
        found = False
        drug_name = input("\nEnter the drug name to search : ")
        print(f'Searching for {drug_name} in the inventory...\n')
        # Open the inventory file
        medicine_inventory = openpyxl.load_workbook("medicine_inventory.xlsx")
        mi = medicine_inventory.active
        for i in range(1, mi.max_row+1):
            for j in range(1, 2):
                med_id = mi.cell(row=i, column=1)
                med_name = mi.cell(row=i, column=2)
                med_qty = mi.cell(row=i, column=4)
                match_drug = re.match(
                    '.*'+drug_name.lower()+'.*', med_name.value.lower())
                if (match_drug):
                    found = True
                    print(f" Medicine Id : {med_id.value}", end=" ")
                    print(f" Medicine Name : {med_name.value}", end=" ")
                    print(f" Quantity Available : {med_qty.value}", end="\n")
        if found is False:
            print("Medicine Not Found !\n")
        while (True):
            try:
                continue_search = str(
                    input("\nDo you wish to search for another medicine? (y/n): "))
                if continue_search == 'y':
                    break
                elif continue_search == 'n':
                    return None
            except ValueError:
                print("Invalid Option !")
