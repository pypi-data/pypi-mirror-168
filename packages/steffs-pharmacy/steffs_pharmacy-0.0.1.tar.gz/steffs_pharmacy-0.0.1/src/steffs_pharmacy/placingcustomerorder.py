
import re
import openpyxl
import time
import tabulate
import clearing
from rich.console import Console
from rich.table import Table


def placing_customer_order(medicine_cart):
    console = Console()
    clearing.clear()
    # Enter the med id
    while True:
        print("\n")
        found_med_flag = False
        qty_available_flag = False
        table_main = Table(show_header=False, header_style="bold blue",
                           title="Placing Customer Order", title_justify="center")
        table_main.add_row(
            'Follow the instructions to add medicines to cart ..')
        console.print(table_main)

        # User input of the med id
        while True:
            try:
                med_id_user_input = int(
                    input("Enter the med id to add to cart.. To discontinue, enter 0 -->  ").strip())
                break
            except ValueError:
                print("Invalid med id !")
                time.sleep(2)
                continue
        # Return to Customer Order Menu
        if med_id_user_input == 0:
            return medicine_cart
        # User input of med quantity
        while True:
            try:
                med_qty_user_input = int(input(
                    "Enter the quantity to add to cart.. To discontinue, enter 0 -->  ").strip())
                break
            except ValueError:
                print("Invalid Quantity !")
                time.sleep(2)
                continue
        # Return to Customer Order Menu
        if med_qty_user_input == 0:
            return medicine_cart
        # Open inventory
        medicine_inventory = openpyxl.load_workbook("MEDICINE_INVENTORY.xlsx")
        mi = medicine_inventory.active
        # Search Inventory for the Medicine and quantity
        for i in range(2, mi.max_row+1):
            # Get the Med id for the row
            med_id = mi.cell(row=i, column=1)
            # Get the Med name for the row
            med_name = mi.cell(row=i, column=2)
            # Get the Med quantity for the row
            med_qty = mi.cell(row=i, column=4)
            # Get the Med price per unit for the row
            med_price = mi.cell(row=i, column=6)
            # Checking if the medicine is available
            if (med_id_user_input == int(med_id.value)):
                # Set the med availability flag to true
                found_med_flag = True
                available_med_qty = med_qty.value
                # Checking if quantity is less than the available quantity
                if (med_qty_user_input <= int(med_qty.value)):
                    # Set quantity availability flag to true
                    qty_available_flag = True
                    # Checking if the medicine is already available in cart.
                    if (next((i for i, x in enumerate(medicine_cart) if x["med_id"] == med_id.value), None) is not None):
                        while True:
                            try:
                                # If the Med is already available in cart, check if it need to be replaced with new entry
                                med_duplicate_user_input = input(
                                    "Med already present in cart. Do you want to replace the existing item in the cart? (y/n)")
                                # If yes then replace the existing item in cart with the new item
                                if (med_duplicate_user_input.lower() == 'y'):
                                    medicine_cart = list(
                                        filter(lambda i: i['med_id'] != med_id.value, medicine_cart))
                                    medicine_cart.append({'med_id': med_id.value, 'med_name': med_name.value,
                                                         'med_qty': med_qty_user_input, 'med_price': med_price.value * med_qty_user_input})
                                    print("Medicine replaced !")
                                    break
                                # if no then skip the step to add to  cart
                                elif (med_duplicate_user_input.lower() == 'n'):
                                    print(
                                        "Operation cancelled . Item not added to cart...")
                                    break
                                else:
                                    # Raise a value error for any invalid inputs.
                                    raise ValueError
                            # Catch any value  error for user inputs
                            except ValueError:
                                print("Invalid Entry (Only y/n) !")
                                time.sleep(1)
                                continue
                    else:
                        # If the medicine is not already in the cart then add it to the cart
                        medicine_cart.append({'med_id': med_id.value, 'med_name': med_name.value,
                                             'med_qty': med_qty_user_input, 'med_price': med_price.value * med_qty_user_input})

                    print("\nMedicine added to cart .. \n")
                    table_cart = Table(show_header=False,
                                       header_style="bold blue", title="Medicine Cart",title_justify="center")
                    header = medicine_cart[0].keys()
                    rows = [x.values() for x in medicine_cart]
                    table_cart.add_row(tabulate.tabulate(rows, header))
                    clearing.clear()
                    console.print(table_cart)

        # if med availability flag is false then provide user alert
        if found_med_flag is False:
            print("Medicine Not Found !")
        # if quantity available in inventory is less than required then provide user alert.
        if found_med_flag is True and qty_available_flag is False:
            print(
                f"There is not enough quantity available for this medicine in the inventory. Only {available_med_qty} available !")
